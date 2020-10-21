from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator

from .forms import TaskForm
from .models import Task
from django.contrib import messages
import datetime

from datetime import timedelta
from openpyxl import Workbook
from django.http import HttpResponse

# Create your views here.

@login_required
def taskList(request):

    search = request.GET.get('search')
    filter = request.GET.get('filter')
    tasksDoneRecently = Task.objects.filter(done='done', updated_at__gt=datetime.datetime.now()-datetime.timedelta(days=30), user=request.user).count()
    tasksDone = Task.objects.filter(done='done', user=request.user).count()
    tasksDoing = Task.objects.filter(done='doing', user=request.user).count()

    if search:
        tasks = Task.objects.filter(title__icontains=search, user=request.user)

    elif filter:
        tasks = Task.objects.filter(done=filter, user=request.user)

    else:

        tasks_list = Task.objects.all().order_by('-created_at').filter(user=request.user)

        paginator = Paginator(tasks_list, 3)
        
        page = request.GET.get('page')

        tasks = paginator.get_page(page)
        
    return render(request, 'tasks/list.html', {'tasks': tasks, 'tasksrecently': tasksDoneRecently, 'tasksdone': tasksDone, 'tasksdoing': tasksDoing})

@login_required
def taskView(request, id):
    task = get_object_or_404(Task, pk=id)
    return render(request, 'tasks/task.html', {'task': task})

@login_required
def newTask(request):
    if request.method == 'POST':
        form = TaskForm(request.POST)

        if form.is_valid():
            task = form.save(commit=False)
            task.done = 'doing'
            task.user = request.user
            task.save()
            return redirect('/')

    else:    
        form = TaskForm()
        return render(request, 'tasks/addtask.html', {'form': form})

@login_required
def editTask(request, id):
    task =get_object_or_404(Task, pk=id)
    form = TaskForm(instance=task)

    if(request.method == 'POST'):
        form = TaskForm(request.POST, instance=task)

        if(form.is_valid()):
            task.save()
            return redirect('/')
        else:
              return render(request, 'tasks/edittask.html', {'form': form, 'task': task})
    else:  
        return render(request, 'tasks/edittask.html', {'form': form, 'task': task})

@login_required
def deleteTask(request, id):
    task =get_object_or_404(Task, pk=id)
    task.delete()

    messages.info(request,'Task deletada com sucesso!')

    return redirect('/')

@login_required
def changeStatus(request,id):
    task =get_object_or_404(Task, pk=id)

    if(task.done == 'doing'):
        task.done = 'done'
    else:
        task.done = 'doing'

    task.save()

    return redirect('/')

def helloWorld(request):
    return HttpResponse('Hello World!')

def yourName(request, name):
    return render(request, 'tasks/yourname.html', {'name': name})

#exportar

@login_required
def export_list(request):
    tasks_queryset= Task.objects.all()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-lista-tarefas.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Tarefas'

    columns = [
        'ID',
        'Tarefa',
        'Descrição',
        'Estado',
        'Data_Criação',
        'Data_Atualização',
    ]
    row_num = 1

# Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

 # Iterate through all movies
    for task in tasks_queryset:
        row_num += 1
        
        # Define the data for each cell in the row 
        row = [
            task.pk,
            task.title,
            task.description,
            task.done,
            task.created_at,
            task.updated_at,
        ]
        
        # Assign the data for each cell of the row 
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response